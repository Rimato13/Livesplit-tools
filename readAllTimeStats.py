from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils import get_column_letter

def getAllTimeStats():
    wb = load_workbook("runData.xlsx")
    sheet = wb.worksheets[0]
    data = []
    i = 1
    while str(sheet['A' + str(i)].value) != 'Optimal':
        y = 1
        fastAmount = 0
        fineAmount = 0
        totalAmount = 0
        while str(sheet[str(get_column_letter(y + 1)) + str(i)].value) != 'None':
            if str(sheet[str(get_column_letter(y + 1)) + str(i)].value) == 'Fast':
                fastAmount += 1
                fineAmount += 1
            elif str(sheet[str(get_column_letter(y + 1)) + str(i)].value) == 'Fine':
                fineAmount += 1
            y += 1
            totalAmount += 1
        data.append({'name': sheet['A' + str(i)].value, 'Fast rate': str("{0:.0%}".format(fastAmount/totalAmount)) , 'Fine rate':str("{0:.0%}".format(fineAmount/totalAmount))})
        i += 1
    df = pd.DataFrame(data)
    print(df.to_string(index=False))

getAllTimeStats()