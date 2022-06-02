from lxml import etree
import xml.etree.ElementTree as ET
from openpyxl import load_workbook, Workbook
from datetime import datetime
import pandas as pd
from openpyxl.utils import get_column_letter
import os.path

standardTimes = []
splitNames = []

def getSplits(firstTag, secondTag):
    parser = etree.XMLParser(strip_cdata=False)
    tree = ET.parse('sms_any%.lss', parser)
    root = tree.getroot()
    latest = ""
    splits = []
    for child in root:
        if (child.tag == "Segments"):
            for child2 in child:
                if (child2.tag == "Segment"):
                    if (latest != "") : 
                        splits.append(latest)
                    for child3 in child2:
                        if (child3.tag == firstTag):
                            latest = ""
                            for child4 in child3:
                                if (child4.tag == secondTag):
                                        for child5 in child4:
                                            if (child5.tag == "RealTime"):
                                                split = child5.text
                                                split = split.rstrip('0').rstrip('.') if '.' in split else split
                                                latest = split
    splits.append(latest)
    return splits    

def getSplitsFromXlsx():
    wb = load_workbook("standards.xlsx")
    sheet = wb.active
    i = 1
    while str(sheet['A' + str(i)].value) != 'None':
        minutes = ""
        seconds = ""
        milliseconds = ""
        splitNames.append(str(sheet['A' + str(i)].value))
        seperatedTimes = str(sheet['B' + str(i)].value).split(":")
        if (len(seperatedTimes) == 1):
            moreSeperatedTimes = seperatedTimes[0].split(".")
            minutes = "00"
            seconds = moreSeperatedTimes[0].zfill(2)
            milliseconds = moreSeperatedTimes[1].zfill(2)
        else:
            moreSeperatedTimes = seperatedTimes[1].split(".")
            minutes = seperatedTimes[0].zfill(2)
            seconds = moreSeperatedTimes[0].zfill(2)
            milliseconds = moreSeperatedTimes[1].zfill(2)
        standardTimes.append(f"00:{minutes}:{seconds}.{milliseconds}")
        i += 1

def saveDataToXlsx(statusList, optimalSplitAmount, consistentSplitAmount):
    if (os.path.isfile('runData.xlsx')) != True:
        wb = Workbook(write_only=True)
        wb.create_sheet(title="Sheet1")
        wb.save("runData.xlsx")
    wb = load_workbook("runData.xlsx")
    sheet = wb.worksheets[0]
    i = 0
    if str(sheet['A1'].value) == 'None': 
        for split in splitNames:
            sheet['A' + str(i + 1)].value = split
            sheet['B' + str(i + 1)].value = statusList[i]
            i += 1
        sheet['A' + str(i + 1)] = "Optimal"
        sheet['B' + str(i + 1)].value = optimalSplitAmount
        sheet['A' + str(i + 2)] = "Fine"
        sheet['B' + str(i + 2)].value = consistentSplitAmount
    else:
        letter = ""
        notEmpty = True
        y = 1
        while notEmpty:
            letter = get_column_letter(y)
            if str(sheet[letter + str(y)].value) == 'None':
                notEmpty = False
            y += 1
        for status in statusList:
            sheet[letter + str(i + 1)].value = status
            i += 1
        sheet[letter + str(i + 1)].value = optimalSplitAmount
        sheet[letter + str(i + 2)].value = consistentSplitAmount
    wb.save("runData.xlsx")

def compareRuns(latestRun, standards):
    optimalStandard = 0
    consistencyStandard = 0
    i = 0
    data = []
    statusList = []
    for split in latestRun:
        splitResult =  compareSplit(split,standards[i])
        status = None
        if(splitResult == "optimal"):
            optimalStandard += 1
            consistencyStandard += 1
            status = "Fast"
        elif (splitResult == "minor_timeloss"):
            consistencyStandard += 1
            status = "Fine"
        else:
            status = "Slow"
        statusList.append(status)
        data.append({'name': splitNames[i], 'split': split, 'standard':standards[i], 'status':status})
        i += 1
    df = pd.DataFrame(data)
    print(df.to_string(index=False))
    print("You met " + str(optimalStandard) + "/" + str(i) + " of the standards which is: " + str("{0:.0%}".format(optimalStandard/i)))
    print("You got close to " + str(consistencyStandard) + "/" + str(i) + " of the standards which is: " + str("{0:.0%}".format(consistencyStandard/i)))
    saveDataToXlsx(statusList,optimalStandard, consistencyStandard)

def compareSplit(split1, split2):
    time_obj = datetime.strptime(split1, '%H:%M:%S.%f' )
    time_obj2 = datetime.strptime(split2, '%H:%M:%S.%f' )
    #test = timedelta(time_obj - time_obj2).total_seconds()
    
    timeComparison = time_obj - time_obj2
    if (timeComparison.days < 0):
        return "optimal"
    timeComparison = datetime.strptime(str(time_obj - time_obj2), '%H:%M:%S.%f' )
    looseStandard = datetime.strptime("00:00:05.01", '%H:%M:%S.%f' )
    if (timeComparison >  looseStandard):
        return "major_timeloss"
    else:
        return "minor_timeloss"

getSplitsFromXlsx()
latestTimes = getSplits("SegmentHistory","Time")
compareRuns(latestTimes, standardTimes)
